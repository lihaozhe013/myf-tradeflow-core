import sqlite3
import openpyxl
from openpyxl.styles import Font
import re
import json
import sys
import argparse
from datetime import datetime
import io
import os

# 修复Windows下print中文json时UnicodeEncodeError，强制stdout为utf-8编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..', 'data.db'))
EXPORT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'exported-files'))
FONT_NAME = '微软雅黑'

TABLES = {
    '1': ('客户/供应商', 'partners'),
    '2': ('产品', 'products'),
    '3': ('产品价格', 'product_prices'),
}

os.makedirs(EXPORT_DIR, exist_ok=True)

def clean_sheet_name(name):
    return re.sub(r'[\\/*?:\[\]]', '-', name)[:31]

def export_base_info(tables=None, output_file=None):
    """
    导出基础信息的核心函数
    :param tables: 要导出的表列表，默认全部
    :param output_file: 输出文件名，默认为时间戳命名
    :return: 包含结果信息的字典
    """
    if tables is None:
        tables = list(TABLES.keys())
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f'base-info-export_{timestamp}.xlsx'
    output_file = os.path.join(EXPORT_DIR, output_file)
    
    try:
        conn = sqlite3.connect(DB_PATH)
        data_dict = {}
        total_records = 0
        
        for t in tables:
            if t not in TABLES:
                continue
            name, table = TABLES[t]
            sql = f"SELECT * FROM {table}"
            cur = conn.execute(sql)
            rows = cur.fetchall()
            headers = [desc[0] for desc in cur.description]
            data_dict[name] = (headers, rows)
            total_records += len(rows)
        
        conn.close()
        
        if data_dict:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            font = Font(name=FONT_NAME)
            
            for sheet_name, (headers, rows) in data_dict.items():
                safe_name = clean_sheet_name(sheet_name)
                ws = wb.create_sheet(title=safe_name)
                ws.append(headers)
                for row in rows:
                    ws.append(row)
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = font
            
            wb.save(output_file)
            
            return {
                'success': True,
                'message': f'导出成功，文件名：{output_file}',
                'file_path': output_file,
                'total_records': total_records,
                'tables_exported': len(data_dict)
            }
        else:
            return {
                'success': False,
                'message': '没有符合条件的数据，无需导出。',
                'file_path': None,
                'total_records': 0,
                'tables_exported': 0
            }
    
    except Exception as e:
        return {
            'success': False,
            'message': f'导出失败：{str(e)}',
            'file_path': None,
            'total_records': 0,
            'tables_exported': 0
        }

def print_table_options():
    print("请选择要导出的基础信息表（可多选，直接输入数字组合，如12导出客户和产品，不填默认全选）：")
    for k, v in TABLES.items():
        print(f"{k}. {v[0]}")
    print("示例：13 只导出客户和产品价格表；全部导出请直接回车")

def get_user_tables():
    print_table_options()
    sel = input("请输入表格编号：").strip()
    if not sel:
        return list(TABLES.keys())
    return [c for c in sel if c in TABLES]

def interactive_mode():
    """交互模式"""
    print("=== 基础信息导出工具 ===")
    tables = get_user_tables()
    result = export_base_info(tables)
    print(f"\n{result['message']}")
    return result

def command_line_mode():
    """命令行模式"""
    parser = argparse.ArgumentParser(description='基础信息导出工具')
    parser.add_argument('--tables', type=str, help='要导出的表格编号，如：123', default='123')
    parser.add_argument('--output', type=str, help='输出文件名', default=None)
    parser.add_argument('--json', action='store_true', help='以JSON格式输出结果')
    
    args = parser.parse_args()
    
    # 解析表格编号
    tables = [c for c in args.tables if c in TABLES]
    
    result = export_base_info(tables, args.output)
    
    if args.json:
        print(json.dumps(result, ensure_ascii=False))
    else:
        print(result['message'])
    
    return result

def main():
    if len(sys.argv) > 1:
        # 命令行模式
        return command_line_mode()
    else:
        # 交互模式
        return interactive_mode()

if __name__ == "__main__":
    main()
