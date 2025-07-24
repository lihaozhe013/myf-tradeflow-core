import sqlite3
import openpyxl
from openpyxl.styles import Font
import re
from datetime import datetime

DB_PATH = '../../data/data.db'
EXPORT_FILE = '应收应付明细导出.xlsx'
FONT_NAME = '微软雅黑'

def clean_sheet_name(name):
    return re.sub(r'[\\/*?:\[\]]', '-', name)[:31]

def fetch_partners(conn, partner_type):
    sql = 'SELECT * FROM partners WHERE type = ?'
    return conn.execute(sql, (partner_type,)).fetchall()

def fetch_receivable_summary(conn, customer_code):
    total_sql = 'SELECT SUM(total_price) FROM outbound_records WHERE customer_code = ?'
    paid_sql = 'SELECT SUM(amount) FROM receivable_payments WHERE customer_code = ?'
    total = conn.execute(total_sql, (customer_code,)).fetchone()[0] or 0
    paid = conn.execute(paid_sql, (customer_code,)).fetchone()[0] or 0
    return total, paid, total - paid

def fetch_payable_summary(conn, supplier_code):
    total_sql = 'SELECT SUM(total_price) FROM inbound_records WHERE supplier_code = ?'
    paid_sql = 'SELECT SUM(amount) FROM payable_payments WHERE supplier_code = ?'
    total = conn.execute(total_sql, (supplier_code,)).fetchone()[0] or 0
    paid = conn.execute(paid_sql, (supplier_code,)).fetchone()[0] or 0
    return total, paid, total - paid

def fetch_outbound(conn, customer_code, date_from=None, date_to=None):
    sql = '''SELECT id, product_code, product_model, quantity, unit_price, total_price, outbound_date, invoice_number, order_number, remark
             FROM outbound_records WHERE customer_code = ?'''
    params = [customer_code]
    if date_from:
        sql += ' AND outbound_date >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND outbound_date <= ?'
        params.append(date_to)
    sql += ' ORDER BY outbound_date DESC'
    return conn.execute(sql, params).fetchall()

def fetch_inbound(conn, supplier_code, date_from=None, date_to=None):
    sql = '''SELECT id, product_code, product_model, quantity, unit_price, total_price, inbound_date, invoice_number, order_number, remark
             FROM inbound_records WHERE supplier_code = ?'''
    params = [supplier_code]
    if date_from:
        sql += ' AND inbound_date >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND inbound_date <= ?'
        params.append(date_to)
    sql += ' ORDER BY inbound_date DESC'
    return conn.execute(sql, params).fetchall()

def fetch_receivable_payments(conn, customer_code, date_from=None, date_to=None):
    sql = '''SELECT id, amount, pay_date, pay_method, remark FROM receivable_payments WHERE customer_code = ?'''
    params = [customer_code]
    if date_from:
        sql += ' AND pay_date >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND pay_date <= ?'
        params.append(date_to)
    sql += ' ORDER BY pay_date DESC'
    return conn.execute(sql, params).fetchall()

def fetch_payable_payments(conn, supplier_code, date_from=None, date_to=None):
    sql = '''SELECT id, amount, pay_date, pay_method, remark FROM payable_payments WHERE supplier_code = ?'''
    params = [supplier_code]
    if date_from:
        sql += ' AND pay_date >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND pay_date <= ?'
        params.append(date_to)
    sql += ' ORDER BY pay_date DESC'
    return conn.execute(sql, params).fetchall()

def write_partner_sheet(wb, partner, summary, outbound_records, payments, is_customer=True):
    name = partner['short_name'] if partner['short_name'] else partner['code']
    sheet_name = clean_sheet_name(name)
    ws = wb.create_sheet(title=sheet_name)
    font = Font(name=FONT_NAME)
    # 基本信息
    ws.append(["简称", "全称", "代号", "地址", "联系人", "电话"])
    ws.append([
        partner['short_name'], partner['full_name'], partner['code'],
        partner['address'], partner['contact_person'], partner['contact_phone']
    ])
    ws.append([])
    # 汇总
    ws.append(["总应收" if is_customer else "总应付", "已回款" if is_customer else "已付款", "余额"])
    ws.append([summary[0], summary[1], summary[2]])
    ws.append([])
    # 出/入库记录
    ws.append(["全部{}记录".format("出库" if is_customer else "入库")])
    ws.append(["单号", "产品代号", "产品型号", "数量", "单价", "总价", "日期", "发票号", "订单号", "备注"])
    for row in outbound_records:
        ws.append(list(row))
    ws.append([])
    # 回/付款记录
    ws.append(["全部{}记录".format("回款" if is_customer else "付款")])
    ws.append(["记录ID", "金额", "日期", "方式", "备注"])
    for row in payments:
        ws.append(list(row))
    # 设置字体
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font

def get_date_input(prompt):
    val = input(f"{prompt}（YYYY-MM-DD，不填为全部）：").strip()
    if val:
        try:
            datetime.strptime(val, "%Y-%m-%d")
            return val
        except ValueError:
            print("日期格式错误，将忽略该筛选。")
    return None

def main():
    print("=== 应收/应付明细导出工具 ===")
    print("可选择出/入库和回/付款的日期区间，不填为全部")
    outbound_from = get_date_input("出库/入库起始日期")
    outbound_to = get_date_input("出库/入库结束日期")
    payment_from = get_date_input("回款/付款起始日期")
    payment_to = get_date_input("回款/付款结束日期")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # 客户（应收）
    customers = fetch_partners(conn, 1)
    for partner in customers:
        summary = fetch_receivable_summary(conn, partner['code'])
        outbound = fetch_outbound(conn, partner['code'], outbound_from, outbound_to)
        payments = fetch_receivable_payments(conn, partner['code'], payment_from, payment_to)
        write_partner_sheet(wb, partner, summary, outbound, payments, is_customer=True)
    # 供应商（应付）
    suppliers = fetch_partners(conn, 0)
    for partner in suppliers:
        summary = fetch_payable_summary(conn, partner['code'])
        inbound = fetch_inbound(conn, partner['code'], outbound_from, outbound_to)
        payments = fetch_payable_payments(conn, partner['code'], payment_from, payment_to)
        write_partner_sheet(wb, partner, summary, inbound, payments, is_customer=False)
    wb.save(EXPORT_FILE)
    print(f"\n✅ 导出成功，文件名：{EXPORT_FILE}")

if __name__ == "__main__":
    main()
