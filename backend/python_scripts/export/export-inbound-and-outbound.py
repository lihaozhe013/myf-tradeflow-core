import sqlite3
import openpyxl
from openpyxl.styles import Font
from tabulate import tabulate
import re

DB_PATH = '../../data/data.db'
EXPORT_FILE = '导出结果.xlsx'
FONT_NAME = '微软雅黑'

TABLES = {
    '1': ('入库记录', 'inbound_records'),
    '2': ('出库记录', 'outbound_records'),
}

def print_table_options():
    print("请选择要导出的表格（可多选，直接输入数字组合，如12导出入库和出库，不填默认全选）：")
    for k, v in TABLES.items():
        print(f"{k}. {v[0]}")
    print("示例：12 只导出入库和出库表；全部导出请直接回车")

def get_user_tables():
    print_table_options()
    sel = input("请输入表格编号：").strip()
    if not sel:
        return list(TABLES.keys())
    return [c for c in sel if c in TABLES]

def get_filter_input(prompt, example):
    val = input(f"{prompt}（不填为全部，示例：{example}）：").strip()
    return val if val else None

def get_filters():
    print("\n可选筛选条件，全部可跳过：")
    date_from = get_filter_input("起始日期(YYYY-MM-DD)", "2024-01-01")
    date_to = get_filter_input("结束日期(YYYY-MM-DD)", "2024-12-31")
    product_code = get_filter_input("相关商品代号", "P001")
    return date_from, date_to, product_code

def build_query(table, date_from, date_to, product_code):
    where = []
    params = []
    if table in ('inbound_records', 'outbound_records'):
        date_col = 'inbound_date' if table == 'inbound_records' else 'outbound_date'
        if date_from:
            where.append(f"{date_col} >= ?")
            params.append(date_from)
        if date_to:
            where.append(f"{date_col} <= ?")
            params.append(date_to)
        if product_code:
            where.append("product_code = ?")
            params.append(product_code)
    sql = f"SELECT * FROM {table}"
    if where:
        sql += " WHERE " + " AND ".join(where)
    return sql, params

def clean_sheet_name(name):
    return re.sub(r'[\\/*?:\[\]]', '-', name)[:31]

def export_to_excel(data_dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    font = Font(name=FONT_NAME)
    for sheet_name, (headers, rows) in data_dict.items():
        safe_name = clean_sheet_name(sheet_name)
        ws = wb.create_sheet(title=safe_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
    wb.save(EXPORT_FILE)
    print(f"\n✅ 导出成功，文件名：{EXPORT_FILE}")

def main():
    print("=== 入库/出库数据导出工具 ===")
    tables = get_user_tables()
    date_from, date_to, product_code = get_filters()
    conn = sqlite3.connect(DB_PATH)
    data_dict = {}
    for t in tables:
        name, table = TABLES[t]
        sql, params = build_query(table, date_from, date_to, product_code)
        cur = conn.execute(sql, params)
        rows = cur.fetchall()
        headers = [desc[0] for desc in cur.description]
        data_dict[name] = (headers, rows)
        print(f"\n【{name}】共导出{len(rows)}条记录。")
        if len(rows) > 0:
            print(tabulate(rows[:5], headers, tablefmt="grid", showindex=True))
            if len(rows) > 5:
                print("...（仅显示前5条）")
    conn.close()
    if data_dict:
        export_to_excel(data_dict)
    else:
        print("没有符合条件的数据，无需导出。")

if __name__ == "__main__":
    main()