import sqlite3
import openpyxl
from openpyxl.styles import Font
import re
import json
import sys
import argparse
from datetime import datetime
import io
import os

DB_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..', 'data.db'))
EXPORT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'exported-files'))
FONT_NAME = '微软雅黑'

# 强制stdout为utf-8编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
os.makedirs(EXPORT_DIR, exist_ok=True)

def clean_sheet_name(name):
    return re.sub(r'[\\/*?:\[\]]', '-', name)[:31]

def fetch_partners(conn, partner_type):
    sql = 'SELECT * FROM partners WHERE type = ?'
    return conn.execute(sql, (partner_type,)).fetchall()

def fetch_receivable_summary(conn, customer_code):
    total_sql = 'SELECT SUM(total_price) FROM outbound_records WHERE customer_code = ?'
    paid_sql = 'SELECT SUM(amount) FROM receivable_payments WHERE customer_code = ?'
    total = conn.execute(total_sql, (customer_code,)).fetchone()[0] or 0
    paid = conn.execute(paid_sql, (customer_code,)).fetchone()[0] or 0
    return total, paid, total - paid

def fetch_payable_summary(conn, supplier_code):
    total_sql = 'SELECT SUM(total_price) FROM inbound_records WHERE supplier_code = ?'
    paid_sql = 'SELECT SUM(amount) FROM payable_payments WHERE supplier_code = ?'
    total = conn.execute(total_sql, (supplier_code,)).fetchone()[0] or 0
    paid = conn.execute(paid_sql, (supplier_code,)).fetchone()[0] or 0
    return total, paid, total - paid

def fetch_outbound(conn, customer_code, date_from=None, date_to=None):
    sql = '''SELECT id, product_code, product_model, quantity, unit_price, total_price, outbound_date, invoice_number, order_number, remark
             FROM outbound_records WHERE customer_code = ?'''
    params = [customer_code]
    if date_from:
        sql += ' AND outbound_date >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND outbound_date <= ?'
        params.append(date_to)
    sql += ' ORDER BY outbound_date DESC'
    return conn.execute(sql, params).fetchall()

def fetch_inbound(conn, supplier_code, date_from=None, date_to=None):
    sql = '''SELECT id, product_code, product_model, quantity, unit_price, total_price, inbound_date, invoice_number, order_number, remark
             FROM inbound_records WHERE supplier_code = ?'''
    params = [supplier_code]
    if date_from:
        sql += ' AND inbound_date >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND inbound_date <= ?'
        params.append(date_to)
    sql += ' ORDER BY inbound_date DESC'
    return conn.execute(sql, params).fetchall()

def fetch_receivable_payments(conn, customer_code, date_from=None, date_to=None):
    sql = '''SELECT id, amount, pay_date, pay_method, remark FROM receivable_payments WHERE customer_code = ?'''
    params = [customer_code]
    if date_from:
        sql += ' AND pay_date >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND pay_date <= ?'
        params.append(date_to)
    sql += ' ORDER BY pay_date DESC'
    return conn.execute(sql, params).fetchall()

def fetch_payable_payments(conn, supplier_code, date_from=None, date_to=None):
    sql = '''SELECT id, amount, pay_date, pay_method, remark FROM payable_payments WHERE supplier_code = ?'''
    params = [supplier_code]
    if date_from:
        sql += ' AND pay_date >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND pay_date <= ?'
        params.append(date_to)
    sql += ' ORDER BY pay_date DESC'
    return conn.execute(sql, params).fetchall()

def write_partner_sheet(wb, partner, summary, outbound_records, payments, is_customer=True):
    name = partner['short_name'] if partner['short_name'] else partner['code']
    sheet_name = clean_sheet_name(name)
    ws = wb.create_sheet(title=sheet_name)
    font = Font(name=FONT_NAME)
    # 基本信息
    ws.append(["简称", "全称", "代号", "地址", "联系人", "电话"])
    ws.append([
        partner['short_name'], partner['full_name'], partner['code'],
        partner['address'], partner['contact_person'], partner['contact_phone']
    ])
    ws.append([])
    # 汇总
    ws.append(["总应收" if is_customer else "总应付", "已回款" if is_customer else "已付款", "余额"])
    ws.append([summary[0], summary[1], summary[2]])
    ws.append([])
    # 出/入库记录
    ws.append(["全部{}记录".format("出库" if is_customer else "入库")])
    ws.append(["单号", "产品代号", "产品型号", "数量", "单价", "总价", "日期", "发票号", "订单号", "备注"])
    for row in outbound_records:
        ws.append(list(row))
    ws.append([])
    # 回/付款记录
    ws.append(["全部{}记录".format("回款" if is_customer else "付款")])
    ws.append(["记录ID", "金额", "日期", "方式", "备注"])
    for row in payments:
        ws.append(list(row))
    # 设置字体
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font

def export_receivable_payable(outbound_from=None, outbound_to=None, payment_from=None, payment_to=None, output_file=None):
    """
    导出应收应付明细的核心函数
    :param outbound_from: 出库/入库起始日期
    :param outbound_to: 出库/入库结束日期
    :param payment_from: 回款/付款起始日期
    :param payment_to: 回款/付款结束日期
    :param output_file: 输出文件名
    :return: 包含结果信息的字典
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f'receivable-payable-export_{timestamp}.xlsx'
    output_file = os.path.join(EXPORT_DIR, output_file)
    
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        total_customers = 0
        total_suppliers = 0
        
        # 客户（应收）
        customers = fetch_partners(conn, 1)
        for partner in customers:
            summary = fetch_receivable_summary(conn, partner['code'])
            outbound = fetch_outbound(conn, partner['code'], outbound_from, outbound_to)
            payments = fetch_receivable_payments(conn, partner['code'], payment_from, payment_to)
            write_partner_sheet(wb, partner, summary, outbound, payments, is_customer=True)
            total_customers += 1
        
        # 供应商（应付）
        suppliers = fetch_partners(conn, 0)
        for partner in suppliers:
            summary = fetch_payable_summary(conn, partner['code'])
            inbound = fetch_inbound(conn, partner['code'], outbound_from, outbound_to)
            payments = fetch_payable_payments(conn, partner['code'], payment_from, payment_to)
            write_partner_sheet(wb, partner, summary, inbound, payments, is_customer=False)
            total_suppliers += 1
        
        conn.close()
        wb.save(output_file)
        
        return {
            'success': True,
            'message': f'导出成功，文件名：{output_file}',
            'file_path': output_file,
            'total_customers': total_customers,
            'total_suppliers': total_suppliers,
            'total_sheets': total_customers + total_suppliers,
            'filters': {
                'outbound_from': outbound_from,
                'outbound_to': outbound_to,
                'payment_from': payment_from,
                'payment_to': payment_to
            }
        }
    
    except Exception as e:
        return {
            'success': False,
            'message': f'导出失败：{str(e)}',
            'file_path': None,
            'total_customers': 0,
            'total_suppliers': 0,
            'total_sheets': 0,
            'filters': {
                'outbound_from': outbound_from,
                'outbound_to': outbound_to,
                'payment_from': payment_from,
                'payment_to': payment_to
            }
        }

def get_date_input(prompt):
    val = input(f"{prompt}（YYYY-MM-DD，不填为全部）：").strip()
    if val:
        try:
            datetime.strptime(val, "%Y-%m-%d")
            return val
        except ValueError:
            print("日期格式错误，将忽略该筛选。")
    return None

def interactive_mode():
    """交互模式"""
    print("=== 应收/应付明细导出工具 ===")
    print("可选择出/入库和回/付款的日期区间，不填为全部")
    outbound_from = get_date_input("出库/入库起始日期")
    outbound_to = get_date_input("出库/入库结束日期")
    payment_from = get_date_input("回款/付款起始日期")
    payment_to = get_date_input("回款/付款结束日期")
    
    result = export_receivable_payable(outbound_from, outbound_to, payment_from, payment_to)
    print(f"\n{result['message']}")
    return result

def command_line_mode():
    """命令行模式"""
    parser = argparse.ArgumentParser(description='应收/应付明细导出工具')
    parser.add_argument('--outbound-from', type=str, help='出库/入库起始日期 (YYYY-MM-DD)', default=None)
    parser.add_argument('--outbound-to', type=str, help='出库/入库结束日期 (YYYY-MM-DD)', default=None)
    parser.add_argument('--payment-from', type=str, help='回款/付款起始日期 (YYYY-MM-DD)', default=None)
    parser.add_argument('--payment-to', type=str, help='回款/付款结束日期 (YYYY-MM-DD)', default=None)
    parser.add_argument('--output', type=str, help='输出文件名', default=None)
    parser.add_argument('--json', action='store_true', help='以JSON格式输出结果')
    
    args = parser.parse_args()
    
    result = export_receivable_payable(args.outbound_from, args.outbound_to, args.payment_from, args.payment_to, args.output)
    
    if args.json:
        print(json.dumps(result, ensure_ascii=False))
    else:
        print(result['message'])
    
    return result

def main():
    if len(sys.argv) > 1:
        # 命令行模式
        return command_line_mode()
    else:
        # 交互模式
        return interactive_mode()

if __name__ == "__main__":
    main()
