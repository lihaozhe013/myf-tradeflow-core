import sqlite3
import openpyxl
from openpyxl.styles import Font
from tabulate import tabulate
import re
import json
import sys
import argparse
from datetime import datetime
import io
import os

# 强制stdout为utf-8编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..', 'data.db'))
FONT_NAME = '微软雅黑'

TABLES = {
    '1': ('入库记录', 'inbound_records'),
    '2': ('出库记录', 'outbound_records'),
}

def clean_sheet_name(name):
    return re.sub(r'[\\/*?:\[\]]', '-', name)[:31]

def build_query(table, date_from, date_to, product_code):
    where = []
    params = []
    if table in ('inbound_records', 'outbound_records'):
        date_col = 'inbound_date' if table == 'inbound_records' else 'outbound_date'
        if date_from:
            where.append(f"{date_col} >= ?")
            params.append(date_from)
        if date_to:
            where.append(f"{date_col} <= ?")
            params.append(date_to)
        if product_code:
            where.append("product_code = ?")
            params.append(product_code)
    sql = f"SELECT * FROM {table}"
    if where:
        sql += " WHERE " + " AND ".join(where)
    return sql, params

def export_inbound_outbound(tables=None, date_from=None, date_to=None, product_code=None, output_file=None):
    """
    导出入库出库记录的核心函数
    :param tables: 要导出的表列表，默认全部
    :param date_from: 起始日期
    :param date_to: 结束日期
    :param product_code: 产品代号
    :param output_file: 输出文件名，默认为时间戳命名
    :return: 包含结果信息的字典
    """
    if tables is None:
        tables = list(TABLES.keys())
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f'入库出库导出_{timestamp}.xlsx'
    
    try:
        conn = sqlite3.connect(DB_PATH)
        data_dict = {}
        total_records = 0
        
        for t in tables:
            if t not in TABLES:
                continue
            name, table = TABLES[t]
            sql, params = build_query(table, date_from, date_to, product_code)
            cur = conn.execute(sql, params)
            rows = cur.fetchall()
            headers = [desc[0] for desc in cur.description]
            data_dict[name] = (headers, rows)
            total_records += len(rows)
        
        conn.close()
        
        if data_dict:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            font = Font(name=FONT_NAME)
            
            for sheet_name, (headers, rows) in data_dict.items():
                safe_name = clean_sheet_name(sheet_name)
                ws = wb.create_sheet(title=safe_name)
                ws.append(headers)
                for row in rows:
                    ws.append(row)
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = font
            
            wb.save(output_file)
            
            return {
                'success': True,
                'message': f'导出成功，文件名：{output_file}',
                'file_path': output_file,
                'total_records': total_records,
                'tables_exported': len(data_dict),
                'filters': {
                    'date_from': date_from,
                    'date_to': date_to,
                    'product_code': product_code
                }
            }
        else:
            return {
                'success': False,
                'message': '没有符合条件的数据，无需导出。',
                'file_path': None,
                'total_records': 0,
                'tables_exported': 0,
                'filters': {
                    'date_from': date_from,
                    'date_to': date_to,
                    'product_code': product_code
                }
            }
    
    except Exception as e:
        return {
            'success': False,
            'message': f'导出失败：{str(e)}',
            'file_path': None,
            'total_records': 0,
            'tables_exported': 0,
            'filters': {
                'date_from': date_from,
                'date_to': date_to,
                'product_code': product_code
            }
        }

def print_table_options():
    print("请选择要导出的表格（可多选，直接输入数字组合，如12导出入库和出库，不填默认全选）：")
    for k, v in TABLES.items():
        print(f"{k}. {v[0]}")
    print("示例：12 只导出入库和出库表；全部导出请直接回车")

def get_user_tables():
    print_table_options()
    sel = input("请输入表格编号：").strip()
    if not sel:
        return list(TABLES.keys())
    return [c for c in sel if c in TABLES]

def get_filter_input(prompt, example):
    val = input(f"{prompt}（不填为全部，示例：{example}）：").strip()
    return val if val else None

def get_filters():
    print("\n可选筛选条件，全部可跳过：")
    date_from = get_filter_input("起始日期(YYYY-MM-DD)", "2024-01-01")
    date_to = get_filter_input("结束日期(YYYY-MM-DD)", "2024-12-31")
    product_code = get_filter_input("相关商品代号", "P001")
    return date_from, date_to, product_code

def interactive_mode():
    """交互模式"""
    print("=== 入库/出库数据导出工具 ===")
    tables = get_user_tables()
    date_from, date_to, product_code = get_filters()
    result = export_inbound_outbound(tables, date_from, date_to, product_code)
    print(f"\n{result['message']}")
    
    # 显示预览
    if result['success']:
        conn = sqlite3.connect(DB_PATH)
        for t in tables:
            if t not in TABLES:
                continue
            name, table = TABLES[t]
            sql, params = build_query(table, date_from, date_to, product_code)
            cur = conn.execute(sql, params)
            rows = cur.fetchall()
            headers = [desc[0] for desc in cur.description]
            print(f"\n【{name}】共导出{len(rows)}条记录。")
            if len(rows) > 0:
                print(tabulate(rows[:5], headers, tablefmt="grid", showindex=True))
                if len(rows) > 5:
                    print("...（仅显示前5条）")
        conn.close()
    
    return result

def command_line_mode():
    """命令行模式"""
    parser = argparse.ArgumentParser(description='入库/出库数据导出工具')
    parser.add_argument('--tables', type=str, help='要导出的表格编号，如：12', default='12')
    parser.add_argument('--date-from', type=str, help='起始日期 (YYYY-MM-DD)', default=None)
    parser.add_argument('--date-to', type=str, help='结束日期 (YYYY-MM-DD)', default=None)
    parser.add_argument('--product-code', type=str, help='产品代号', default=None)
    parser.add_argument('--output', type=str, help='输出文件名', default=None)
    parser.add_argument('--json', action='store_true', help='以JSON格式输出结果')
    
    args = parser.parse_args()
    
    # 解析表格编号
    tables = [c for c in args.tables if c in TABLES]
    
    result = export_inbound_outbound(tables, args.date_from, args.date_to, args.product_code, args.output)
    
    if args.json:
        print(json.dumps(result, ensure_ascii=False))
    else:
        print(result['message'])
    
    return result

def main():
    if len(sys.argv) > 1:
        # 命令行模式
        return command_line_mode()
    else:
        # 交互模式
        return interactive_mode()

if __name__ == "__main__":
    main()
