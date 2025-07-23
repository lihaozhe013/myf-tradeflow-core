import sqlite3
import openpyxl
from openpyxl.styles import Font
import re

DB_PATH = '../../data/data.db'
EXPORT_FILE = '基础信息导出.xlsx'
FONT_NAME = '微软雅黑'

TABLES = {
    '1': ('客户/供应商', 'partners'),
    '2': ('产品', 'products'),
    '3': ('产品价格', 'product_prices'),
}

def print_table_options():
    print("请选择要导出的基础信息表（可多选，直接输入数字组合，如12导出客户和产品，不填默认全选）：")
    for k, v in TABLES.items():
        print(f"{k}. {v[0]}")
    print("示例：13 只导出客户和产品价格表；全部导出请直接回车")

def get_user_tables():
    print_table_options()
    sel = input("请输入表格编号：").strip()
    if not sel:
        return list(TABLES.keys())
    return [c for c in sel if c in TABLES]

def clean_sheet_name(name):
    return re.sub(r'[\\/*?:\[\]]', '-', name)[:31]

def export_to_excel(data_dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    font = Font(name=FONT_NAME)
    for sheet_name, (headers, rows) in data_dict.items():
        safe_name = clean_sheet_name(sheet_name)
        ws = wb.create_sheet(title=safe_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
    wb.save(EXPORT_FILE)
    print(f"\n✅ 导出成功，文件名：{EXPORT_FILE}")

def main():
    print("=== 基础信息导出工具 ===")
    tables = get_user_tables()
    conn = sqlite3.connect(DB_PATH)
    data_dict = {}
    for t in tables:
        name, table = TABLES[t]
        sql = f"SELECT * FROM {table}"
        cur = conn.execute(sql)
        rows = cur.fetchall()
        headers = [desc[0] for desc in cur.description]
        data_dict[name] = (headers, rows)
        print(f"\n【{name}】共导出{len(rows)}条记录。")
    conn.close()
    if data_dict:
        export_to_excel(data_dict)
    else:
        print("没有符合条件的数据，无需导出。")

if __name__ == "__main__":
    main()

